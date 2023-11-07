import os
import openpyxl
from datetime import datetime

#Nesarios de instalacion:
#openpyxl

# , fecha, modificado, ult_acceso
def escribeExcel(path, nombre, ruta, peso, autor):
    libro_excel = openpyxl.load_workbook(path)
    sheet = libro_excel.active
    sheet.title = 'Excels'
    fila = sheet.max_row + 1
    
    sheet.cell(row=fila, column=1, value=nombre)
    sheet.cell(row=fila, column=2, value=ruta)
    sheet.cell(row=fila, column=3, value=peso)
    sheet.cell(row=fila, column=4, value=autor)

    libro_excel.save(path)
    libro_excel.close



    


# Ubicandonos en carpeta origen
ruta = os.getcwd()
ruta_modificada = ruta.replace('\\', '/') + '/inventario_carpetas'
print(f'la ruta normal: {ruta}')
print(f'ruta modificada: {ruta_modificada}')

path = os.getcwd()
modified_path = path.replace('\\', '/') + '/origen'

os.chdir(ruta_modificada)
print(f'actualmente en: {os.getcwd()}')

ruta_documento_final_excel = ruta_modificada + '/Inventario_Carpetas.xlsx'

#Creando Excel  y encabezados
libro_excel = openpyxl.Workbook()

#Escribiendo encabezados
sheet = libro_excel.active # <- Activa la hoja a trabajar
filas = 1
columnas = 1
#Escribe en las posiciones dadas
sheet.cell(row=filas, column=columnas, value= 'NOMBRE')
columnas += 1
sheet.cell(row=filas, column=columnas, value= 'RUTA')
columnas += 1
sheet.cell(row=filas, column=columnas, value= 'PESO')
columnas += 1
sheet.cell(row=filas, column=columnas, value= 'AUTOR')
columnas += 1
sheet.cell(row=filas, column=columnas, value= 'FECHA CREACION')
columnas += 1
sheet.cell(row=filas, column=columnas, value= 'MODIFICADO')
columnas += 1
sheet.cell(row=filas, column=columnas, value= 'ACCESADO POR ULTIMA VEZ')


libro_excel.save('Inventario_Carpetas.xlsx')
libro_excel.close()

#Empieza lo bonito

os.chdir(modified_path)
print(f'\n\npath actual: {modified_path}\n\n\n\n')           #( Origen )

# Rastreando archivox excel

col = 1
fil = 2

for archivoExcel in os.listdir():
    #os.chdir(modified_path) #( Origen )

    nombre, extension = os.path.splitext(archivoExcel)
    print(f'revisando que es {archivoExcel}')
    if extension == '.xlsx':
        
        #Recopilando datos    
        excel = openpyxl.load_workbook(archivoExcel)
        propiedades = libro_excel.properties
        nombre_excel = archivoExcel
        autor = propiedades.creator
        ruta_excel = os.getcwd()
        tamanio_excel = os.path.getsize(archivoExcel)
        tamano_archivo_mb = tamanio_excel / (1024 * 1024)
        print(f'tamano en MB: {tamano_archivo_mb}')
        fechaCreacion_excel = propiedades.created
        fechaModificacion = propiedades.modified
        ultimo_acceso = os.path.getatime(archivoExcel)
        fecha_ultimo_acceso = datetime.fromtimestamp(ultimo_acceso)
        excel.close()
        escribeExcel(ruta_documento_final_excel, nombre_excel, ruta_excel,  tamano_archivo_mb, autor)



        #Escribiendo esos datos en el excel:
        #print(f'col: {col} fila: {fil}')

        #sheet.cell(row=fil, column=col, value= nombre_excel)
        #fil += 1
        #print('Si hizo lo de arriba xd?')
        #excel.save('Inventario_Carpetas.xlsx')
        #excel.close()















